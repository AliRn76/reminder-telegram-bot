from datetime import datetime
from typing import Iterator

import jdatetime
import openpyxl
import requests

from configs import BOT_TOKEN, CHAT_ID, BIRTHDAY_MESSAGE, ANNIVERSARY_MESSAGE


def read_excel() -> Iterator[tuple[str, datetime.date, datetime.date]]:
    dataframe = openpyxl.load_workbook('users.xlsx').active

    for row in range(0, dataframe.max_row):
        first_name, last_name, birthday, anniversary_day = dataframe.iter_cols(1, 4)
        if first_name[row].value:
            birthday = jdatetime.date(*map(int, birthday[row].value.split('/'))).togregorian()
            anniversary_day = jdatetime.date(*map(int, anniversary_day[row].value.split('/'))).togregorian()
            yield first_name[row].value, last_name[row].value, birthday, anniversary_day


def find_birthdays() -> Iterator[str]:
    today = datetime.now()

    for first_name, _, birthday, _ in read_excel():
        if today.month == birthday.month and today.day == birthday.day:
            yield first_name


def find_anniversaries() -> Iterator[str]:
    today = datetime.now()

    for first_name, last_name, _, anniversary in read_excel():
        if today.month == anniversary.month and today.day == anniversary.day:
            yield f'{first_name} {last_name}'


def send_message(message):
    url = f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage'
    try:
        requests.post(url, json={'chat_id': CHAT_ID, 'text': message})
    except Exception as e:
        print(e)


def send_birthday_messages(name):
    message = BIRTHDAY_MESSAGE.format(name)
    send_message(message=message)


def send_anniversary_messages(name):
    message = ANNIVERSARY_MESSAGE.format(name)
    send_message(message=message)
